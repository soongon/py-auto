import openpyxl

# 워크북 (엑셀파일) 오픈
wb = openpyxl.load_workbook('./data/movies.xlsx')

print(type(wb))
print(wb)

# 작업 할 워크시트 오픈
sheets = wb.sheetnames
print(sheets)

sheet = wb['Sheet1']
print(sheet)

# A1, B2 .. 이런 형태로 셀에 접근
cell = sheet['D5']
print(cell.value)

print(sheet.max_row)
print(sheet.max_column)

wb.close()
