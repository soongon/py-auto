import openpyxl

# 워크북 오픈
wb = openpyxl.load_workbook('./data/movies.xlsx')
# 워크시트 오픈
sheet = wb.active

# 셀에 접근하여 작업.. cell() 함수 사용
print(sheet.cell(2, 3).value)

print(sheet.cell(row=10, column=4).value)
print('=======================================')
# 감독만 출력
for i in range(2, sheet.max_row + 1):
    print(sheet.cell(i, 4).value)

wb.close()
