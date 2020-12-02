import openpyxl
from openpyxl.comments import Comment

wb = openpyxl.load_workbook('./data/movies.xlsx')
sheet = wb['Sheet1']
sheet.title = '영화랭킹'

# 11 을 만든다. (쓰기)
sheet.cell(12, 1).value = 11

# 바람과함께 사라지다에 코멘트 달기
sheet.cell(2, 3).comment = Comment('바람과 함께 사라지다...', author='김순곤')

# 파일로 저장..
wb.save('./data/movies_updated.xlsx')
print('job completed..')